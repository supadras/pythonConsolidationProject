import io

import numpy as np
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import pandas as pd
from datetime import datetime
from io import StringIO
import os
from openpyxl import workbook, load_workbook

#writer = pd.ExcelWriter('ConsolidatedScores.xlsx', mode="a", engine='openpyxl')
#writer.book = load_workbook('ConsolidatedScores.xlsx')
#writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
#if startrow is None and sheet_name in writer.book.sheetnames:
#   startrow = writer.book[sheet_name].max_row

gauth = GoogleAuth()
gauth.LocalWebserverAuth()# client_secrets.json need to be in the same directory as the script
drive = GoogleDrive(gauth)
consolidatedEngScorecsv = pd.DataFrame(columns=['id'])
consolidatedProfScorecsv = pd.DataFrame(columns=['id'])
consolidatedActScorecsv = pd.DataFrame(columns=['id'])
consolidatedLeadScorecsv = pd.DataFrame(columns=['id'])
fileList = drive.ListFile({'q': "'1gUTv3PqgjHzj-HWK6F-1n4awNwx6kW-h' in parents and trashed=false"}).GetList()
listdf = pd.DataFrame(data=fileList, columns=['title', 'id', 'downloadUrl'])
listdf['DateRun']=listdf.title.apply(lambda item: datetime.strptime(str(str(item).split('_')[-1]).replace('.csv', ''), '%m%d%Y').date())
listdf = listdf.sort_values('DateRun')
i = 1
#To Do - include code to remove executed files
for index, file in listdf.iterrows():
    if not(str(file.title).__eq__('Lead_Scoring_03142022.csv')):
        gfile = drive.CreateFile({'id': file.id})
        download_path = os.path.expanduser('~/data')
        try:
            os.makedirs(download_path)
        except FileExistsError:
            pass
        output_fname = os.path.join(download_path, file.title)
        print("Processing file:" + str(i) + file.title)
        #if output_fname is None:
        #    output_fname = file.id
        contents = gfile.GetContentFile(output_fname)
        currentfile = pd.read_csv(output_fname)
        currentfile.set_index('id', inplace=True)
        if i == 1:
            # onetime activity
            consolidatedEngScorecsv.id = currentfile.index
            consolidatedEngScorecsv.set_index('id', inplace=True)
            consolidatedEngScorecsv['R'+str(i)] = currentfile.engagement_score.values
            consolidatedProfScorecsv.id = currentfile.index
            consolidatedProfScorecsv.set_index('id', inplace=True)
            consolidatedProfScorecsv['R' + str(i)] = currentfile.profile_complete_score.values
            consolidatedActScorecsv.id = currentfile.index
            consolidatedActScorecsv.set_index('id', inplace=True)
            consolidatedActScorecsv['R' + str(i)] = currentfile.activity_score.values
            consolidatedLeadScorecsv.id = currentfile.index
            consolidatedLeadScorecsv.set_index('id', inplace=True)
            consolidatedLeadScorecsv['R' + str(i)] = currentfile.lead_score.values
        else:
            listCommonIDs=consolidatedEngScorecsv.index.intersection(currentfile.index.values)
            newdf = currentfile[~(currentfile.index.isin(listCommonIDs))]
            #consolidatedEngScorecsv.append(pd.DataFrame({name: newdf.index, 'R1': np.zeros(len(newdf)), 'R'+str(i): newdf['engagement_score']}))
            testEngdf = pd.DataFrame({'R1': np.zeros(len(newdf)), 'R' + str(i): newdf['engagement_score']},
                                     index=newdf.index)  # profile_complete_score
            testProfdf = pd.DataFrame({'R1': np.zeros(len(newdf)), 'R' + str(i): newdf['profile_complete_score']},
                                     index=newdf.index)  # profile_complete_score
            testActdf = pd.DataFrame({'R1': np.zeros(len(newdf)), 'R' + str(i): newdf['activity_score']},
                                     index=newdf.index)  # profile_complete_score
            testLeaddf = pd.DataFrame({'R1': np.zeros(len(newdf)), 'R' + str(i): newdf['lead_score']}, index=newdf.index) #profile_complete_score
            #testdf.name = newdf.index
            consolidatedEngScorecsv.insert(len(consolidatedEngScorecsv.columns), 'R'+str(i), np.zeros(len(consolidatedEngScorecsv)))
            consolidatedProfScorecsv.insert(len(consolidatedProfScorecsv.columns), 'R' + str(i),
                                           np.zeros(len(consolidatedProfScorecsv)))
            consolidatedActScorecsv.insert(len(consolidatedActScorecsv.columns), 'R' + str(i),
                                           np.zeros(len(consolidatedActScorecsv)))
            consolidatedLeadScorecsv.insert(len(consolidatedLeadScorecsv.columns), 'R' + str(i),
                                           np.zeros(len(consolidatedLeadScorecsv)))
            consolidatedEngScorecsv = pd.concat([consolidatedEngScorecsv, testEngdf])
            consolidatedProfScorecsv = pd.concat([consolidatedProfScorecsv, testProfdf])
            consolidatedActScorecsv = pd.concat([consolidatedActScorecsv, testActdf])
            consolidatedLeadScorecsv = pd.concat([consolidatedLeadScorecsv, testLeaddf])
            #consolidatedEngScorecsv.loc[list(newdf.id.values), 'R'+str(i)] = newdf['engagement_score'].values
            #testdf = pd.dataFrame({'id': newdf['id'], 'R1': np.zeros(len(newdf)), 'R'+str(i): newdf['engagement_score']})
            consolidatedEngScorecsv.loc[listCommonIDs.values, 'R' + str(i)] = currentfile.loc[listCommonIDs.values, 'engagement_score'].values
            consolidatedProfScorecsv.loc[listCommonIDs.values, 'R' + str(i)] = currentfile.loc[
                listCommonIDs.values, 'profile_complete_score'].values
            consolidatedActScorecsv.loc[listCommonIDs.values, 'R' + str(i)] = currentfile.loc[
                listCommonIDs.values, 'activity_score'].values
            consolidatedLeadScorecsv.loc[listCommonIDs.values, 'R' + str(i)] = currentfile.loc[
                listCommonIDs.values, 'lead_score'].values

            #consolidatedEngScorecsv.iloc[listCommonIDs, 'R' + str(i)] = currentfile.iloc[listCommonIDs,'engagement_score']
            #for index, row in currentfile.iterrows():
            #   if (consolidatedEngScorecsv['id'].equals(row.id)):
            #       consolidatedEngScorecsv[len(consolidatedEngScorecsv)] = row.id
            #   consolidatedEngScorecsv['R' + str(i)] = row.engagement_score.values
    i = i+1
    print("Processed file:"+str(i-1)+file.title)
sheet_name = 'EngagementScore'#'ProfileCompletenessScore'
with pd.ExcelWriter('ConsolidatedScores.xlsx') as writer:
    #writer.book = load_workbook('ConsolidatedScores.xlsx')
    #writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    #if startrow is None and sheet_name in writer.book.sheetnames:
    #    startrow = writer.book[sheet_name].max_row
    #consolidatedEngScorecsv.to_excel(writer, sheet_name='EngagementScore')
    consolidatedEngScorecsv.to_excel(writer,   sheet_name='EngagementScore')
    consolidatedProfScorecsv.to_excel(writer, sheet_name='ProfileCompletenessScore')
    consolidatedActScorecsv.to_excel(writer, sheet_name='ActivityScore')
    consolidatedLeadScorecsv.to_excel(writer, sheet_name='LeadScore')

    writer.save()
listdf.to_csv("ExecutedFiles.csv")

def drive_read_csv(url):
    res = requests.get(url, headers={"Authorization": "Bearer " + gauth.attr['credentials'].access_token})
    values = pd.read_csv(BytesIO(res.content))
    return values




